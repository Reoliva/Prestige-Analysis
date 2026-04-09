#9-04-2026
from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import networkx as nx
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from scipy.stats import spearmanr
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import RandomForestRegressor
from sklearn.impute import SimpleImputer
from sklearn.inspection import permutation_importance
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import GroupKFold
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder


DEFAULT_INPUT = "Consolidado_Ordenado.xlsx"
DEFAULT_OUTPUT = "Reporte_Prestigio.xlsx"

COLS = {
    "id": "ID",
    "group": "Group",
    "role": "Rol",
    "team_member": "Team Member",
    "gender": "Gender",
    "speak_a": "Speaking Time Activity A",
    "speak_b": "Speaking Time Activity B",
    "first_a": "First Speaker Activity A (0 No - 1 Si)",
    "first_b": "First Speaker Activity B (0 No - 1 Si)",
    "course_isf": "Cursó ISF (0 No - 1 Si)",
    "course_fa": "Cursó FA (0 No - 1 Si)",
    "acad": "Normaliced Nota Promedio (0-1)",
    "exp_proj": "Normaliced en proyectos de software reales (0-1)",
    "exp_plan": "Normaliced Experiencia con Planning Poker o Planning Game (0-1)",
    "exp_model": "Normaliced Experiencia en modelado conceptual de software (0-1)",
    "peer": "Normaliced Promedio total Desempeño percibido Sin Autoevaluación (0-1)",
}

RE_TOTAL_ATTN = re.compile(
    r"^Total Visual Attention Received from the member Activity ([A-Z]) TM(\d+)$",
    flags=re.IGNORECASE,
)
RE_SPEAK_ATTN = re.compile(
    r"^Visual attention when speaking received from the member Activity ([A-Z]) TM(\d+)$",
    flags=re.IGNORECASE,
)
RE_SILENT_ATTN = re.compile(
    r"^Visual attention when not speaking received from the member Activity ([A-Z]) TM(\d+)$",
    flags=re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Prestige analysis pipeline with data preservation, QA checks, corrected normalization, group-aware ML, and plot-ready outputs."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Input Excel file.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output Excel file.")
    parser.add_argument(
        "--formula-step",
        type=float,
        default=0.1,
        help="Step size for prestige formula brute-force search. Default: 0.1",
    )
    return parser.parse_args()


def normalize_colname(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip())


def load_input(path: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Load the workbook and return both a raw-preserved copy and a cleaned copy."""
    df_raw = pd.read_excel(path)
    df_raw.columns = [normalize_colname(c) for c in df_raw.columns]

    df = df_raw.copy()

    protected_text_cols = {COLS["role"], COLS["gender"]}

    for col in df.columns:
        if col in protected_text_cols:
            df[col] = (
                df[col]
                .astype("string")
                .str.strip()
                .str.replace(r"\s+", " ", regex=True)
            )
            continue

        if pd.api.types.is_numeric_dtype(df[col]):
            continue

        original = df[col].copy()
        cleaned_str = (
            original.astype("string")
            .str.strip()
            .replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "NaN": pd.NA})
        )

        parsed = pd.to_numeric(
            cleaned_str.str.replace(",", ".", regex=False),
            errors="coerce",
        )

        non_empty = cleaned_str.notna().sum()
        parsed_count = parsed.notna().sum()

        if non_empty > 0 and parsed_count / non_empty >= 0.85:
            df[col] = parsed
        else:
            df[col] = cleaned_str

    if COLS["role"] in df.columns:
        df["Rol_Clean"] = df[COLS["role"]].astype("string").str.strip().str.replace(r"\s+", " ", regex=True)
    if COLS["gender"] in df.columns:
        df["Gender_Clean"] = df[COLS["gender"]].astype("string").str.strip().str.upper()

    return df_raw, df


def validate_required_columns(df: pd.DataFrame) -> List[str]:
    return [col for col in COLS.values() if col not in df.columns]


def minmax_01(series: pd.Series) -> pd.Series:
    series = pd.to_numeric(series, errors="coerce")
    valid = series.dropna()
    if valid.empty:
        return pd.Series(np.nan, index=series.index, dtype="float64")
    min_val = valid.min()
    max_val = valid.max()
    if pd.isna(min_val) or pd.isna(max_val) or math.isclose(max_val, min_val):
        return pd.Series(0.0, index=series.index, dtype="float64")
    return (series - min_val) / (max_val - min_val)


def zscore_within_group(series: pd.Series) -> pd.Series:
    series = pd.to_numeric(series, errors="coerce")
    valid = series.dropna()
    if valid.empty:
        return pd.Series(np.nan, index=series.index, dtype="float64")
    std = valid.std(ddof=0)
    if pd.isna(std) or math.isclose(std, 0.0):
        return pd.Series(0.0, index=series.index, dtype="float64")
    return (series - valid.mean()) / std


def find_matching_columns(columns: Iterable[str], pattern: re.Pattern) -> List[str]:
    return sorted([col for col in columns if pattern.match(col)])


def sum_columns(df: pd.DataFrame, columns: Sequence[str]) -> pd.Series:
    if not columns:
        return pd.Series(0.0, index=df.index, dtype="float64")
    return df[list(columns)].apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1)


def build_attention_graph_metrics(df: pd.DataFrame, total_attention_cols: Sequence[str]) -> pd.DataFrame:
    rows: List[Dict[str, float]] = []

    for group_id, group_df in df.groupby(COLS["group"], dropna=True):
        group_df = group_df.copy()
        graph = nx.DiGraph()

        members = group_df[COLS["team_member"]].tolist()
        graph.add_nodes_from(members)

        for _, row in group_df.iterrows():
            target = row[COLS["team_member"]]
            for col in total_attention_cols:
                match = RE_TOTAL_ATTN.match(col)
                if not match:
                    continue
                source = int(match.group(2))
                val = pd.to_numeric(row[col], errors="coerce")
                weight = 0.0 if pd.isna(val) else float(val)
                if source == target or weight <= 0:
                    continue
                if graph.has_edge(source, target):
                    graph[source][target]["weight"] += weight
                else:
                    graph.add_edge(source, target, weight=weight)

        try:
            eigen = nx.eigenvector_centrality(graph, weight="weight", max_iter=5000)
        except Exception:
            eigen = {node: np.nan for node in graph.nodes}

        try:
            pagerank = nx.pagerank(graph, weight="weight")
        except Exception:
            pagerank = {node: np.nan for node in graph.nodes}

        in_degree = dict(graph.in_degree(weight="weight"))
        out_degree = dict(graph.out_degree(weight="weight"))

        for _, row in group_df.iterrows():
            tm = row[COLS["team_member"]]
            rows.append(
                {
                    COLS["id"]: row[COLS["id"]],
                    "Group": group_id,
                    "Team Member": tm,
                    "SNA_Eigenvector": eigen.get(tm, np.nan),
                    "SNA_PageRank": pagerank.get(tm, np.nan),
                    "SNA_InDegree_Weighted": in_degree.get(tm, 0.0),
                    "SNA_OutDegree_Weighted": out_degree.get(tm, 0.0),
                    "SNA_Attention_Balance": in_degree.get(tm, 0.0) - out_degree.get(tm, 0.0),
                }
            )

    return pd.DataFrame(rows)


def compute_derived_features(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, List[str]]]:
    df_out = df.copy()

    total_attention_cols = find_matching_columns(df_out.columns, RE_TOTAL_ATTN)
    speaking_attention_cols = find_matching_columns(df_out.columns, RE_SPEAK_ATTN)
    silent_attention_cols = find_matching_columns(df_out.columns, RE_SILENT_ATTN)
    speaking_time_cols = [c for c in [COLS["speak_a"], COLS["speak_b"]] if c in df_out.columns]

    base_metric_columns = [
        COLS["acad"],
        COLS["exp_proj"],
        COLS["exp_plan"],
        COLS["exp_model"],
        COLS["peer"],
    ]
    base_metric_columns = [c for c in base_metric_columns if c in df_out.columns]

    if COLS["acad"] in df_out.columns:
        df_out["Prest_Abs_Acad_Raw"] = pd.to_numeric(df_out[COLS["acad"]], errors="coerce")
        df_out["Prest_Abs_Acad_01"] = minmax_01(df_out[COLS["acad"]])

    exp_raw_cols = [c for c in [COLS["exp_proj"], COLS["exp_plan"], COLS["exp_model"]] if c in df_out.columns]
    if exp_raw_cols:
        df_out["Prest_Abs_Exp_Raw"] = df_out[exp_raw_cols].apply(pd.to_numeric, errors="coerce").mean(axis=1)
        scaled_exp = [minmax_01(df_out[c]).rename(f"{c}__01") for c in exp_raw_cols]
        df_out["Prest_Abs_Exp_01"] = pd.concat(scaled_exp, axis=1).mean(axis=1)

    if COLS["peer"] in df_out.columns:
        df_out["Prest_Abs_Peer_Raw"] = pd.to_numeric(df_out[COLS["peer"]], errors="coerce")
        df_out["Prest_Abs_Peer_01"] = minmax_01(df_out[COLS["peer"]])

    for source, target in [
        ("Prest_Abs_Acad_01", "Prest_Rel_Acad"),
        ("Prest_Abs_Exp_01", "Prest_Rel_Exp"),
        ("Prest_Abs_Peer_01", "Prest_Rel_Peer"),
    ]:
        if source in df_out.columns:
            df_out[target] = df_out.groupby(COLS["group"], dropna=False)[source].transform(zscore_within_group)

    df_out["Total_Attention_Received"] = sum_columns(df_out, total_attention_cols)
    df_out["Gaze_While_Speaking"] = sum_columns(df_out, speaking_attention_cols)
    df_out["Gaze_While_Silent"] = sum_columns(df_out, silent_attention_cols)
    df_out["Total_Speaking_Time"] = sum_columns(df_out, speaking_time_cols)
    if COLS["first_a"] in df_out.columns and COLS["first_b"] in df_out.columns:
        df_out["First_Speaker_Count"] = (
            pd.to_numeric(df_out[COLS["first_a"]], errors="coerce").fillna(0)
            + pd.to_numeric(df_out[COLS["first_b"]], errors="coerce").fillna(0)
        )

    for source, target in [
        ("Total_Attention_Received", "Attention_Received_Share_Group"),
        ("Gaze_While_Speaking", "Gaze_Speaking_Share_Group"),
        ("Gaze_While_Silent", "Gaze_Silent_Share_Group"),
        ("Total_Speaking_Time", "Speaking_Time_Share_Group"),
    ]:
        if source in df_out.columns:
            totals = df_out.groupby(COLS["group"], dropna=False)[source].transform("sum")
            df_out[target] = np.where(totals != 0, df_out[source] / totals, np.nan)

    graph_metrics = build_attention_graph_metrics(df_out, total_attention_cols)
    if not graph_metrics.empty:
        df_out = df_out.merge(graph_metrics, on=[COLS["id"], "Group", "Team Member"], how="left")

    metadata = {
        "total_attention_cols": total_attention_cols,
        "speaking_attention_cols": speaking_attention_cols,
        "silent_attention_cols": silent_attention_cols,
        "speaking_time_cols": speaking_time_cols,
        "base_metric_columns": base_metric_columns,
    }

    qa_rows = []
    for col in base_metric_columns:
        numeric = pd.to_numeric(df_out[col], errors="coerce")
        qa_rows.append(
            {
                "Column": col,
                "Min_Original": numeric.min(),
                "Max_Original": numeric.max(),
                "Rows_Out_Expected_0_1": int(((numeric < 0) | (numeric > 1)).fillna(False).sum()),
                "Rows_Missing": int(numeric.isna().sum()),
            }
        )
    qa_df = pd.DataFrame(qa_rows)

    return df_out, qa_df, metadata


def spearman_table(df: pd.DataFrame, x_cols: Sequence[str], y_cols: Sequence[str], family: str) -> pd.DataFrame:
    rows = []
    for x_col in x_cols:
        if x_col not in df.columns:
            continue
        for y_col in y_cols:
            if y_col not in df.columns:
                continue
            pair = df[[x_col, y_col]].apply(pd.to_numeric, errors="coerce").dropna()
            if len(pair) < 3:
                rho = np.nan
                p_val = np.nan
                n_obs = len(pair)
            else:
                rho, p_val = spearmanr(pair[x_col], pair[y_col])
                n_obs = len(pair)
            rows.append(
                {
                    "family": family,
                    "x_var": x_col,
                    "y_var": y_col,
                    "rho": rho,
                    "p_value": p_val,
                    "n": n_obs,
                    "abs_rho": abs(rho) if pd.notna(rho) else np.nan,
                }
            )
    return pd.DataFrame(rows).sort_values(["family", "abs_rho"], ascending=[True, False]).reset_index(drop=True)


def generate_weight_grid(step: float) -> List[Tuple[float, float, float]]:
    if step <= 0 or step > 1:
        raise ValueError("formula_step must be > 0 and <= 1")

    step_str = f"{step:.10f}".rstrip("0")
    decimals = len(step_str.split(".")[1]) if "." in step_str else 0
    points = np.round(np.arange(0, 1 + step / 2, step), decimals)
    weights = []
    for w_a in points:
        for w_e in points:
            w_p = round(1 - w_a - w_e, decimals)
            if w_p < -10 ** (-decimals - 2):
                continue
            if w_p < 0:
                w_p = 0.0
            if math.isclose(w_a + w_e + w_p, 1.0, abs_tol=max(1e-9, step / 10)):
                weights.append((round(float(w_a), decimals), round(float(w_e), decimals), round(float(w_p), decimals)))
    return sorted(set(weights))


def run_formula_search(df: pd.DataFrame, step: float) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    components = {
        "Acad": "Prest_Abs_Acad_01",
        "Exp": "Prest_Abs_Exp_01",
        "Peer": "Prest_Abs_Peer_01",
    }
    targets = [
        "Gaze_While_Speaking",
        "Gaze_While_Silent",
        "Total_Speaking_Time",
        "Total_Attention_Received",
        "SNA_Eigenvector",
    ]
    targets = [t for t in targets if t in df.columns]

    for required in components.values():
        if required not in df.columns:
            raise ValueError(f"Missing required prestige component for formula search: {required}")

    formula_values = {}
    dict_rows = []

    for w_a, w_e, w_p in generate_weight_grid(step):
        formula_name = f"Mix_A{int(round(w_a * 100)):02d}_E{int(round(w_e * 100)):02d}_P{int(round(w_p * 100)):02d}"
        values = (
            w_a * df[components["Acad"]]
            + w_e * df[components["Exp"]]
            + w_p * df[components["Peer"]]
        )
        formula_values[formula_name] = values
        dict_rows.append(
            {
                "Formula_Name": formula_name,
                "Formula_Text": f"{w_a:.2f}*Acad + {w_e:.2f}*Exp + {w_p:.2f}*Peer",
                "w_Acad": w_a,
                "w_Exp": w_e,
                "w_Peer": w_p,
                "Component_Acad": components["Acad"],
                "Component_Exp": components["Exp"],
                "Component_Peer": components["Peer"],
            }
        )

    formulas_df = pd.DataFrame(formula_values)

    ranking_rows = []
    for formula_name in formulas_df.columns:
        for target in targets:
            pair = pd.concat([formulas_df[formula_name], df[target]], axis=1).dropna()
            pair.columns = ["formula_value", "target_value"]
            if len(pair) < 3:
                rho = np.nan
                p_val = np.nan
                n_obs = len(pair)
            else:
                rho, p_val = spearmanr(pair["formula_value"], pair["target_value"])
                n_obs = len(pair)
            ranking_rows.append(
                {
                    "Formula_Name": formula_name,
                    "Target": target,
                    "Spearman_rho": rho,
                    "p_value": p_val,
                    "n": n_obs,
                    "abs_rho": abs(rho) if pd.notna(rho) else np.nan,
                }
            )

    ranking_df = pd.DataFrame(ranking_rows)
    formula_dict_df = pd.DataFrame(dict_rows)
    ranking_df = ranking_df.merge(formula_dict_df, on="Formula_Name", how="left")

    best_rows = []
    for target in targets:
        subset = ranking_df[ranking_df["Target"] == target].sort_values(by=["abs_rho", "p_value"], ascending=[False, True])
        if not subset.empty:
            best_rows.append(subset.iloc[0].to_dict())
    best_df = pd.DataFrame(best_rows)

    return best_df, ranking_df.sort_values(["Target", "abs_rho"], ascending=[True, False]), formula_dict_df


def get_model_features(df: pd.DataFrame) -> Tuple[List[str], List[str], List[str]]:
    numeric_features = [
        "Prest_Abs_Acad_01",
        "Prest_Abs_Exp_01",
        "Prest_Abs_Peer_01",
        "Prest_Rel_Acad",
        "Prest_Rel_Exp",
        "Prest_Rel_Peer",
        COLS["course_isf"],
        COLS["course_fa"],
        "First_Speaker_Count",
    ]
    categorical_features = ["Gender_Clean", "Rol_Clean"]

    numeric_features = [c for c in numeric_features if c in df.columns]
    categorical_features = [c for c in categorical_features if c in df.columns]
    return numeric_features + categorical_features, numeric_features, categorical_features


def build_ml_pipeline(numeric_features: Sequence[str], categorical_features: Sequence[str]) -> Pipeline:
    preprocess = ColumnTransformer(
        transformers=[
            (
                "num",
                Pipeline(steps=[("imputer", SimpleImputer(strategy="median"))]),
                list(numeric_features),
            ),
            (
                "cat",
                Pipeline(
                    steps=[
                        ("imputer", SimpleImputer(strategy="most_frequent")),
                        ("onehot", OneHotEncoder(handle_unknown="ignore")),
                    ]
                ),
                list(categorical_features),
            ),
        ],
        remainder="drop",
    )

    model = RandomForestRegressor(
        n_estimators=150,
        random_state=42,
        min_samples_leaf=2,
        n_jobs=1,
    )

    return Pipeline(steps=[("preprocess", preprocess), ("model", model)])


def run_grouped_ml(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    targets = [
        "Gaze_While_Speaking",
        "Gaze_While_Silent",
        "Total_Speaking_Time",
        "Total_Attention_Received",
        "SNA_Eigenvector",
    ]
    targets = [t for t in targets if t in df.columns]

    all_features, numeric_features, categorical_features = get_model_features(df)

    metrics_rows = []
    perm_rows = []
    full_impurity_rows = []

    if not all_features or not targets:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    for target in targets:
        model_df = df[[COLS["group"]] + all_features + [target]].copy()
        model_df = model_df.dropna(subset=[target]).reset_index(drop=True)
        if len(model_df) < 10:
            continue

        X = model_df[all_features]
        y = pd.to_numeric(model_df[target], errors="coerce")
        groups = model_df[COLS["group"]]

        n_groups = groups.nunique()
        if n_groups < 3:
            continue

        n_splits = min(5, n_groups)
        gkf = GroupKFold(n_splits=n_splits)
        pipe = build_ml_pipeline(numeric_features, categorical_features)

        fold_perm = []
        for fold_idx, (train_idx, test_idx) in enumerate(gkf.split(X, y, groups=groups), start=1):
            X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
            y_train, y_test = y.iloc[train_idx], y.iloc[test_idx]

            pipe.fit(X_train, y_train)
            preds = pipe.predict(X_test)

            metrics_rows.append(
                {
                    "Target": target,
                    "Fold": fold_idx,
                    "Train_Rows": len(train_idx),
                    "Test_Rows": len(test_idx),
                    "Train_Groups": groups.iloc[train_idx].nunique(),
                    "Test_Groups": groups.iloc[test_idx].nunique(),
                    "R2": r2_score(y_test, preds),
                    "MAE": mean_absolute_error(y_test, preds),
                    "RMSE": float(np.sqrt(mean_squared_error(y_test, preds))),
                }
            )

            perm = permutation_importance(
                pipe,
                X_test,
                y_test,
                n_repeats=8,
                random_state=42,
                scoring="neg_mean_absolute_error",
                n_jobs=1,
            )
            fold_perm_df = pd.DataFrame(
                {
                    "Target": target,
                    "Fold": fold_idx,
                    "Feature": all_features,
                    "Permutation_Importance_Mean": perm.importances_mean,
                    "Permutation_Importance_Std": perm.importances_std,
                }
            )
            fold_perm.append(fold_perm_df)

        if fold_perm:
            perm_target_df = pd.concat(fold_perm, ignore_index=True)
            perm_summary = (
                perm_target_df.groupby(["Target", "Feature"], as_index=False)
                .agg(
                    Permutation_Importance_Mean=("Permutation_Importance_Mean", "mean"),
                    Permutation_Importance_Std=("Permutation_Importance_Mean", "std"),
                    Mean_Fold_Std=("Permutation_Importance_Std", "mean"),
                )
                .fillna(0)
            )
            perm_summary["Rank"] = perm_summary.groupby("Target")["Permutation_Importance_Mean"].rank(ascending=False, method="dense")
            perm_rows.append(perm_summary)

        pipe.fit(X, y)
        preprocess = pipe.named_steps["preprocess"]
        feature_names = list(preprocess.get_feature_names_out())
        model = pipe.named_steps["model"]
        impurity_df = pd.DataFrame(
            {
                "Target": target,
                "Transformed_Feature": feature_names,
                "Impurity_Importance": model.feature_importances_,
            }
        ).sort_values("Impurity_Importance", ascending=False)
        impurity_df["Rank"] = np.arange(1, len(impurity_df) + 1)
        full_impurity_rows.append(impurity_df)

    metrics_df = pd.DataFrame(metrics_rows)
    if not metrics_df.empty:
        summary = (
            metrics_df.groupby("Target", as_index=False)
            .agg(
                Fold_Count=("Fold", "count"),
                Mean_R2=("R2", "mean"),
                Std_R2=("R2", "std"),
                Mean_MAE=("MAE", "mean"),
                Std_MAE=("MAE", "std"),
                Mean_RMSE=("RMSE", "mean"),
                Std_RMSE=("RMSE", "std"),
            )
            .fillna(0)
        )
        separator = pd.DataFrame([{col: np.nan for col in metrics_df.columns}])
        metrics_df = pd.concat([metrics_df.sort_values(["Target", "Fold"]).reset_index(drop=True), separator, summary], ignore_index=True)

    perm_df = pd.concat(perm_rows, ignore_index=True) if perm_rows else pd.DataFrame()
    impurity_df = pd.concat(full_impurity_rows, ignore_index=True) if full_impurity_rows else pd.DataFrame()
    return metrics_df, perm_df, impurity_df


def build_scatter_pairs(df: pd.DataFrame, corr_long: pd.DataFrame, top_n: int = 10) -> pd.DataFrame:
    if corr_long.empty:
        return pd.DataFrame()

    top_pairs = corr_long.sort_values(["abs_rho", "p_value"], ascending=[False, True]).dropna(subset=["rho"]).head(top_n).copy()
    rows = []
    base_id_cols = [c for c in [COLS["id"], COLS["group"], COLS["team_member"], "Gender_Clean", "Rol_Clean"] if c in df.columns]
    for idx, pair in top_pairs.iterrows():
        pair_id = f"P{idx+1:02d}"
        x_var = pair["x_var"]
        y_var = pair["y_var"]
        pair_df = df[base_id_cols + [x_var, y_var]].copy()
        pair_df["pair_id"] = pair_id
        pair_df["family"] = pair["family"]
        pair_df["x_var"] = x_var
        pair_df["y_var"] = y_var
        pair_df["rho"] = pair["rho"]
        pair_df["p_value"] = pair["p_value"]
        pair_df["x_value"] = pd.to_numeric(pair_df[x_var], errors="coerce")
        pair_df["y_value"] = pd.to_numeric(pair_df[y_var], errors="coerce")
        rows.append(pair_df.drop(columns=[x_var, y_var]))

    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()


def autosize_and_style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    accent_fill = PatternFill("solid", fgColor="EAF4E2")

    for ws in wb.worksheets:
        if ws.max_row == 0 or ws.max_column == 0:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if ws.max_row >= 2 and ws.title in {"00_README", "04_QA_Ranges", "10_ML_CV_Metrics"}:
            for cell in ws[2]:
                cell.fill = subheader_fill

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

        if ws.title == "10_ML_CV_Metrics":
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                first_val = row[0].value
                if isinstance(first_val, str) and first_val == "Fold_Count":
                    for cell in row:
                        cell.fill = accent_fill

    wb.save(path)


def write_excel_report(
    output_path: Path,
    readme_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    clean_df: pd.DataFrame,
    derived_df: pd.DataFrame,
    qa_df: pd.DataFrame,
    corr_abs_df: pd.DataFrame,
    corr_rel_df: pd.DataFrame,
    best_formulas_df: pd.DataFrame,
    formula_ranking_df: pd.DataFrame,
    formula_dict_df: pd.DataFrame,
    ml_metrics_df: pd.DataFrame,
    ml_perm_df: pd.DataFrame,
    ml_impurity_df: pd.DataFrame,
    heatmap_long_df: pd.DataFrame,
    scatter_pairs_df: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="00_README", index=False)
        raw_df.to_excel(writer, sheet_name="01_Raw_Input", index=False)
        clean_df.to_excel(writer, sheet_name="02_Cleaned_Input", index=False)
        derived_df.to_excel(writer, sheet_name="03_Derived_Data", index=False)
        qa_df.to_excel(writer, sheet_name="04_QA_Ranges", index=False)
        corr_abs_df.to_excel(writer, sheet_name="05_Corr_Abs", index=False)
        corr_rel_df.to_excel(writer, sheet_name="06_Corr_Rel", index=False)
        best_formulas_df.to_excel(writer, sheet_name="07_Formula_Best", index=False)
        formula_ranking_df.to_excel(writer, sheet_name="08_Formula_Ranking", index=False)
        formula_dict_df.to_excel(writer, sheet_name="09_Formula_Dict", index=False)
        ml_metrics_df.to_excel(writer, sheet_name="10_ML_CV_Metrics", index=False)
        ml_perm_df.to_excel(writer, sheet_name="11_ML_Permutation", index=False)
        ml_impurity_df.to_excel(writer, sheet_name="12_ML_Impurity", index=False)
        heatmap_long_df.to_excel(writer, sheet_name="13_Heatmap_Long", index=False)
        scatter_pairs_df.to_excel(writer, sheet_name="14_Scatter_Pairs", index=False)

    autosize_and_style_workbook(output_path)


def try_render_workbook(path: Path) -> None:
    try:
        from artifact_tool import SpreadsheetArtifact
        artifact = SpreadsheetArtifact.load(str(path))
        artifact.render()
    except Exception as exc:
        print(f"[Aviso] No se pudo renderizar con artifact_tool: {exc}")


def build_readme(input_path: Path, output_path: Path, formula_step: float, metadata: Dict[str, List[str]], missing_columns: Sequence[str]) -> pd.DataFrame:
    rows = [
        {"Section": "Input", "Detail": str(input_path.name)},
        {"Section": "Output", "Detail": str(output_path.name)},
        {"Section": "Formula_Step", "Detail": formula_step},
        {"Section": "Preservation", "Detail": "Se conserva una copia cruda del input y una versión limpia antes de agregar variables derivadas."},
        {"Section": "Normalization_Fix", "Detail": "Las columnas de prestigio originales se preservan y además se agregan versiones corregidas en escala [0,1] con sufijo _01."},
        {"Section": "ML_Fix", "Detail": "Gender y Rol se mantienen como categóricas y se codifican mediante OneHotEncoder dentro de un pipeline."},
        {"Section": "Validation_Fix", "Detail": "La evaluación ML usa GroupKFold por Group para reducir fuga de información entre miembros del mismo equipo."},
        {"Section": "Missing_Required_Columns", "Detail": ", ".join(missing_columns) if missing_columns else "Ninguna"},
    ]
    for key, value in metadata.items():
        rows.append({"Section": key, "Detail": "; ".join(value) if value else "(sin columnas detectadas)"})
    return pd.DataFrame(rows)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"No se encontró el archivo de entrada: {input_path}")

    raw_df, clean_df = load_input(input_path)
    missing_columns = validate_required_columns(clean_df)
    if missing_columns:
        raise ValueError(f"Faltan columnas requeridas en el input: {missing_columns}")

    derived_df, qa_df, metadata = compute_derived_features(clean_df)

    abs_cols = [c for c in ["Prest_Abs_Acad_01", "Prest_Abs_Exp_01", "Prest_Abs_Peer_01"] if c in derived_df.columns]
    rel_cols = [c for c in ["Prest_Rel_Acad", "Prest_Rel_Exp", "Prest_Rel_Peer"] if c in derived_df.columns]
    behavior_cols = [
        c for c in [
            "Gaze_While_Speaking",
            "Gaze_While_Silent",
            "Total_Speaking_Time",
            "Total_Attention_Received",
            "SNA_Eigenvector",
            "Speaking_Time_Share_Group",
            "Attention_Received_Share_Group",
        ] if c in derived_df.columns
    ]

    corr_abs_df = spearman_table(derived_df, abs_cols, behavior_cols, family="absolute_scaled")
    corr_rel_df = spearman_table(derived_df, rel_cols, behavior_cols, family="relative_group_zscore")
    best_formulas_df, formula_ranking_df, formula_dict_df = run_formula_search(derived_df, args.formula_step)
    ml_metrics_df, ml_perm_df, ml_impurity_df = run_grouped_ml(derived_df)

    heatmap_long_df = pd.concat([corr_abs_df, corr_rel_df], ignore_index=True)
    scatter_pairs_df = build_scatter_pairs(derived_df, heatmap_long_df, top_n=12)

    readme_df = build_readme(input_path, output_path, args.formula_step, metadata, missing_columns)

    write_excel_report(
        output_path=output_path,
        readme_df=readme_df,
        raw_df=raw_df,
        clean_df=clean_df,
        derived_df=derived_df,
        qa_df=qa_df,
        corr_abs_df=corr_abs_df,
        corr_rel_df=corr_rel_df,
        best_formulas_df=best_formulas_df,
        formula_ranking_df=formula_ranking_df,
        formula_dict_df=formula_dict_df,
        ml_metrics_df=ml_metrics_df,
        ml_perm_df=ml_perm_df,
        ml_impurity_df=ml_impurity_df,
        heatmap_long_df=heatmap_long_df,
        scatter_pairs_df=scatter_pairs_df,
    )

    print(f"OK -> {output_path}")


if __name__ == "__main__":
    main()
